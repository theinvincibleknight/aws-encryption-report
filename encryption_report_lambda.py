"""
AWS Encryption Artifact Collector - Lambda Function
Collects EC2 (EBS Volume), RDS, and S3 encryption details across multiple AWS accounts
and all regions using STS cross-account role assumption. Outputs an Excel report to S3.
"""

import boto3
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from io import BytesIO

# Cross-account role ARNs
ACCOUNTS = {
    'dev': {
        'role_arn': 'arn:aws:iam::111111111111:role/inventory-readonly-role',
    },
    'uat': {
        'role_arn': 'arn:aws:iam::222222222222:role/inventory-readonly-role',
    },
    'prod': {
        'role_arn': 'arn:aws:iam::333333333333:role/inventory-readonly-role',
    },
    'oldprod': {
        'role_arn': 'arn:aws:iam::444444444444:role/inventory-readonly-role',
    },
    'network': {
        'role_arn': 'arn:aws:iam::555555555555:role/inventory-readonly-role',
    },
    'sharedservice': {
        'role_arn': 'arn:aws:iam::666666666666:role/inventory-readonly-role',
    },
}

OUTPUT_BUCKET = 'aws-artifact-collector'


def get_all_regions(session):
    """Get all enabled AWS regions for the account."""
    ec2_client = session.client('ec2', region_name='ap-south-1')
    regions = ec2_client.describe_regions(
        Filters=[{'Name': 'opt-in-status', 'Values': ['opt-in-not-required', 'opted-in']}]
    )
    return [region['RegionName'] for region in regions['Regions']]


def assume_role(role_arn, session_name='EncryptionAuditSession'):
    """Assume a cross-account IAM role and return temporary credentials."""
    sts_client = boto3.client('sts')
    response = sts_client.assume_role(
        RoleArn=role_arn,
        RoleSessionName=session_name,
        DurationSeconds=3600
    )
    credentials = response['Credentials']
    session = boto3.Session(
        aws_access_key_id=credentials['AccessKeyId'],
        aws_secret_access_key=credentials['SecretAccessKey'],
        aws_session_token=credentials['SessionToken'],
    )
    return session


def get_ec2_encryption_details(session, account_name, region):
    """Collect EBS volume encryption details for all EC2 instances in a region."""
    ec2_client = session.client('ec2', region_name=region)
    results = []

    # Get all instances
    paginator = ec2_client.get_paginator('describe_instances')
    instances = []
    for page in paginator.paginate():
        for reservation in page['Reservations']:
            for instance in reservation['Instances']:
                instances.append(instance)

    if not instances:
        return results

    # Get all volumes
    vol_paginator = ec2_client.get_paginator('describe_volumes')
    volumes = {}
    for page in vol_paginator.paginate():
        for volume in page['Volumes']:
            volumes[volume['VolumeId']] = volume

    for instance in instances:
        instance_id = instance['InstanceId']
        # Get instance name from tags
        instance_name = instance_id
        if 'Tags' in instance:
            for tag in instance['Tags']:
                if tag['Key'] == 'Name':
                    instance_name = tag['Value']
                    break

        # Check each attached volume
        for block_device in instance.get('BlockDeviceMappings', []):
            volume_id = block_device.get('Ebs', {}).get('VolumeId')
            if volume_id and volume_id in volumes:
                volume = volumes[volume_id]
                encrypted = volume.get('Encrypted', False)
                kms_key_id = volume.get('KmsKeyId', '')

                if encrypted:
                    encryption_type = get_kms_alias(session, kms_key_id, region) if kms_key_id else 'Encrypted (key unknown)'
                else:
                    encryption_type = 'Not Encrypted'

                results.append({
                    'account_name': account_name.upper(),
                    'resource_name': instance_name,
                    'region': region,
                    'encryption_type': encryption_type,
                })

    return results


def get_s3_encryption_details(session, account_name):
    """Collect S3 bucket encryption details (S3 is global, called once per account)."""
    s3_client = session.client('s3', region_name='us-east-1')
    results = []

    try:
        buckets = s3_client.list_buckets().get('Buckets', [])
    except Exception as e:
        print(f"Error listing S3 buckets for {account_name}: {e}")
        return results

    for bucket in buckets:
        bucket_name = bucket['Name']
        encryption_type = 'Not Encrypted'

        try:
            enc_response = s3_client.get_bucket_encryption(Bucket=bucket_name)
            rules = enc_response.get('ServerSideEncryptionConfiguration', {}).get('Rules', [])
            if rules:
                sse_algorithm = rules[0].get('ApplyServerSideEncryptionByDefault', {}).get('SSEAlgorithm', '')
                kms_key = rules[0].get('ApplyServerSideEncryptionByDefault', {}).get('KMSMasterKeyID', '')

                if sse_algorithm == 'aws:kms':
                    encryption_type = get_kms_alias(session, kms_key, 'us-east-1') if kms_key else 'KMS (key unknown)'
                elif sse_algorithm == 'AES256':
                    encryption_type = 'Server-side encryption with Amazon S3 managed keys (SSE-S3)'
                elif sse_algorithm == 'aws:kms:dsse':
                    encryption_type = 'Dual-layer server-side encryption with AWS KMS keys (DSSE-KMS)'
                else:
                    encryption_type = sse_algorithm
        except s3_client.exceptions.ClientError as e:
            if e.response['Error']['Code'] == 'ServerSideEncryptionConfigurationNotFoundError':
                encryption_type = 'Not Encrypted'
            else:
                encryption_type = f'Error: {e.response["Error"]["Code"]}'
        except Exception as e:
            encryption_type = f'Error: {str(e)}'

        results.append({
            'account_name': account_name.upper(),
            'resource_name': bucket_name,
            'encryption_type': encryption_type,
        })

    return results


def get_rds_encryption_details(session, account_name, region):
    """Collect RDS instance encryption details in a region."""
    rds_client = session.client('rds', region_name=region)
    results = []

    try:
        paginator = rds_client.get_paginator('describe_db_instances')
        for page in paginator.paginate():
            for db_instance in page['DBInstances']:
                db_identifier = db_instance['DBInstanceIdentifier']
                endpoint = db_instance.get('Endpoint', {}).get('Address', 'N/A')
                encrypted = db_instance.get('StorageEncrypted', False)
                kms_key_id = db_instance.get('KmsKeyId', '')

                if encrypted:
                    encryption_type = get_kms_alias(session, kms_key_id, region) if kms_key_id else 'Encrypted (key unknown)'
                else:
                    encryption_type = 'Not Encrypted'

                results.append({
                    'account_name': account_name.upper(),
                    'resource_name': db_identifier,
                    'endpoint': endpoint,
                    'region': region,
                    'encryption_type': encryption_type,
                })
    except Exception as e:
        print(f"Error describing RDS instances for {account_name} in {region}: {e}")

    return results


def get_kms_alias(session, kms_key_id, region):
    """Resolve a KMS key ARN/ID to a friendly alias name."""
    kms_client = session.client('kms', region_name=region)
    try:
        # Extract key ID from ARN if needed
        key_id = kms_key_id.split('/')[-1] if '/' in kms_key_id else kms_key_id

        aliases = kms_client.list_aliases(KeyId=key_id).get('Aliases', [])
        if aliases:
            alias_name = aliases[0]['AliasName'].replace('alias/', '')
            return f'KMS: {alias_name}'
        return f'KMS: {key_id}'
    except Exception:
        # Fallback: try to extract alias from ARN pattern
        if 'alias/' in kms_key_id:
            return f'KMS: {kms_key_id.split("alias/")[-1]}'
        return f'KMS: (custom key)'


def style_worksheet(ws, headers):
    """Apply formatting to a worksheet."""
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 60)


def create_excel_report(ec2_data, s3_data, rds_data):
    """Create an Excel workbook with EC2, S3, and RDS sheets."""
    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- EC2 Sheet ---
    ws_ec2 = wb.active
    ws_ec2.title = 'EC2'
    ec2_headers = ['Sr No.', 'Account Name', 'Resource Name', 'Region', 'Encryption Type']
    style_worksheet(ws_ec2, ec2_headers)

    for idx, row in enumerate(ec2_data, 1):
        ws_ec2.append([idx, row['account_name'], row['resource_name'], row['region'], row['encryption_type']])

    for row in ws_ec2.iter_rows(min_row=2, max_row=ws_ec2.max_row, max_col=len(ec2_headers)):
        for cell in row:
            cell.border = thin_border

    # --- S3 Sheet ---
    ws_s3 = wb.create_sheet('S3')
    s3_headers = ['Sr No.', 'Account Name', 'Resource Name', 'Encryption Type']
    style_worksheet(ws_s3, s3_headers)

    for idx, row in enumerate(s3_data, 1):
        ws_s3.append([idx, row['account_name'], row['resource_name'], row['encryption_type']])

    for row in ws_s3.iter_rows(min_row=2, max_row=ws_s3.max_row, max_col=len(s3_headers)):
        for cell in row:
            cell.border = thin_border

    # --- RDS Sheet ---
    ws_rds = wb.create_sheet('RDS')
    rds_headers = ['Sr No.', 'Account Name', 'Resource Name', 'Endpoint', 'Region', 'Encryption']
    style_worksheet(ws_rds, rds_headers)

    for idx, row in enumerate(rds_data, 1):
        ws_rds.append([idx, row['account_name'], row['resource_name'], row['endpoint'], row['region'], row['encryption_type']])

    for row in ws_rds.iter_rows(min_row=2, max_row=ws_rds.max_row, max_col=len(rds_headers)):
        for cell in row:
            cell.border = thin_border

    return wb


def upload_to_s3(workbook):
    """Save workbook to S3 in the required path structure."""
    now = datetime.utcnow()
    year = now.strftime('%Y')
    month = now.strftime('%m')
    filename = f'encryption_report_{year}_{month}.xlsx'
    s3_key = f'encryption/{year}/{month}/{filename}'

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    s3_client = boto3.client('s3')
    s3_client.put_object(
        Bucket=OUTPUT_BUCKET,
        Key=s3_key,
        Body=buffer.getvalue(),
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    print(f"Report uploaded to s3://{OUTPUT_BUCKET}/{s3_key}")
    return f's3://{OUTPUT_BUCKET}/{s3_key}'


def lambda_handler(event, context):
    """Main Lambda entry point."""
    print("Starting encryption artifact collection...")

    all_ec2_data = []
    all_s3_data = []
    all_rds_data = []

    for account_name, account_config in ACCOUNTS.items():
        role_arn = account_config['role_arn']
        print(f"Processing account: {account_name} ({role_arn})")

        try:
            session = assume_role(role_arn)

            # Get all enabled regions for this account
            regions = get_all_regions(session)
            print(f"  Found {len(regions)} regions to scan")

            # Collect S3 encryption (global service, only once per account)
            print(f"  Collecting S3 encryption for {account_name}...")
            s3_data = get_s3_encryption_details(session, account_name)
            all_s3_data.extend(s3_data)
            print(f"    Found {len(s3_data)} S3 bucket entries")

            # Collect EC2 and RDS per region
            for region in regions:
                # EC2 (EBS) encryption
                ec2_data = get_ec2_encryption_details(session, account_name, region)
                if ec2_data:
                    all_ec2_data.extend(ec2_data)
                    print(f"    [{region}] Found {len(ec2_data)} EC2 volume entries")

                # RDS encryption
                rds_data = get_rds_encryption_details(session, account_name, region)
                if rds_data:
                    all_rds_data.extend(rds_data)
                    print(f"    [{region}] Found {len(rds_data)} RDS instance entries")

        except Exception as e:
            print(f"  ERROR processing account {account_name}: {e}")
            continue

    # Generate Excel report
    print("Generating Excel report...")
    workbook = create_excel_report(all_ec2_data, all_s3_data, all_rds_data)

    # Upload to S3
    s3_path = upload_to_s3(workbook)

    summary = {
        'statusCode': 200,
        'body': {
            'message': 'Encryption report generated successfully',
            'report_location': s3_path,
            'summary': {
                'ec2_volumes': len(all_ec2_data),
                's3_buckets': len(all_s3_data),
                'rds_instances': len(all_rds_data),
            }
        }
    }
    print(f"Completed: {summary}")
    return summary
